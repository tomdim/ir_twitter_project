import xlrd
import config
import nltk
import traceback
from openpyxl import load_workbook

pos_lex = []
neg_lex = []
category = ["TweetsTsipras", "TweetsMitsotakis", "TweetsSyriza", "TweetsNeaDimokratia"]

#week = 'Week 2'
#_date = datetime.date.today() - timedelta(days=0)
#_date = _date.strftime("%Y-%m-%d").strip()

class Results:
    positive_tsipras = 0
    negative_tsipras = 0
    neutral_tsipras = 0
    positive_mitsotakis = 0
    negative_mitsotakis = 0
    neutral_mitsotakis = 0
    positive_syriza = 0
    negative_syriza = 0
    neutral_syriza = 0
    positive_nd = 0
    negative_nd = 0
    neutral_nd = 0

class Classifier:
    def parseLex(self, f):
        lex = []
        workbook = xlrd.open_workbook(f, 'utf-8')
        worksheet = workbook.sheet_by_index(0)
        for i in range(1, worksheet.nrows):
            lex.append(worksheet.cell(i,2).value)
            print worksheet.cell(i,2).value
            
        return lex
    
    def getTweets(self, _date, _week):
        res = Results()
        for c in category:
            print "______________________________________________"
            print c, _date
            print "______________________________________________"
            path = 'tweets/%s/%s.txt' % (c,_date)
            tweets = []
            for line in open(path, 'r'):
                tweets.append(line.decode('utf-8'))
                print line
            for i in range(0, len(tweets)):
                self.classifyTweet(tweets[i], c, res)
                print "______________________________________________"
            
        self.storeResults(res, _week, _date)
        
    def classifyTweet(self, text, category, res):
        print "Classifying text:", text, "of category:", category
        pos_words = []
        neg_words = []
        t = nltk.word_tokenize(text)
        try:          
            for i in range(0, len(t)):
                if t[i] in pos_lex:
                    pos_words.append(t[i])
                elif t[i] in neg_lex:
                    neg_words.append(t[i])
        except:
            traceback.print_exc()
        
        self.writeDownResults(category, text, pos_words, neg_words, res)
        
        
        
    def writeDownResults(self, category, text, pos_words, neg_words, res): 
        if "tsipras" in category.lower():
            if len(pos_words) > len(neg_words):
                print "SENTIMENT: Positive"
                res.positive_tsipras += 1
            elif len(pos_words) < len(neg_words):
                print "SENTIMENT: Negative"
                res.negative_tsipras += 1
            else:
                print "SENTIMENT: Neutral"
                res.neutral_tsipras += 1
        elif "mitsotakis" in category.lower():
            if len(pos_words) > len(neg_words):
                print "SENTIMENT: Positive"
                res.positive_mitsotakis += 1
            elif len(pos_words) < len(neg_words):
                print "SENTIMENT: Negative"
                res.negative_mitsotakis += 1
            else:
                print "SENTIMENT: Neutral"
                res.neutral_mitsotakis += 1
        elif "syriza" in category.lower():
            if len(pos_words) > len(neg_words):
                print "SENTIMENT: Positive"
                res.positive_syriza += 1
            elif len(pos_words) < len(neg_words):
                print "SENTIMENT: Negative"
                res.negative_syriza += 1
            else:
                print "SENTIMENT: Neutral"
                res.neutral_syriza += 1
        elif "neadimokratia" in category.lower():
            if len(pos_words) > len(neg_words):
                print "SENTIMENT: Positive"
                res.positive_nd += 1
            elif len(pos_words) < len(neg_words):
                print "SENTIMENT: Negative"
                res.negative_nd += 1
            else:
                print "SENTIMENT: Neutral"
                res.neutral_nd += 1
        else:
            print "Something is wrong here!"
            
        print "\n***POSITIVE WORDS***"
        for i in range(0, len(pos_words)):
            print pos_words[i]
        print "\n***NEGATIVE WORDS***"
        for i in range(0, len(neg_words)):
            print neg_words[i]
        
    def storeResults(self, res, _week, _date):     
        workbook = load_workbook('results.xlsx')
        worksheet = workbook.get_sheet_by_name(_week)
        r = worksheet.max_row + 1
        print r
        worksheet.cell(row = r, column = 1).value = _date
        worksheet.cell(row = r, column = 2).value = res.positive_tsipras
        worksheet.cell(row = r, column = 3).value = res.negative_tsipras
        worksheet.cell(row = r, column = 4).value = res.neutral_tsipras
        worksheet.cell(row = r, column = 5).value = res.positive_mitsotakis
        worksheet.cell(row = r, column = 6).value = res.negative_mitsotakis
        worksheet.cell(row = r, column = 7).value = res.neutral_mitsotakis
        worksheet.cell(row = r, column = 8).value = res.positive_syriza
        worksheet.cell(row = r, column = 9).value = res.negative_syriza
        worksheet.cell(row = r, column = 10).value = res.neutral_syriza
        worksheet.cell(row = r, column = 11).value = res.positive_nd
        worksheet.cell(row = r, column = 12).value = res.negative_nd
        worksheet.cell(row = r, column = 13).value = res.neutral_nd
        
        workbook.save('results.xlsx')
        print "Results of the day stored in: results.xlsx"
        
if __name__ == '__main__':    
    c = Classifier()
    pos_lex = c.parseLex('PosLex.xls')
    print "*********************************************************************"
    neg_lex = c.parseLex('NegLex.xls')

    #---------------------------------------------------
    #Week 1: 2017-01-02 to 2017-01-08
    #Week 2: 2017-01-09 to 2017-01-15
    #format of dates: Year-Month-Day
    _week = 'Week 1'
    _date = '2017-01-02'
    #---------------------------------------------------
    c.getTweets(_date, _week)
    
    
